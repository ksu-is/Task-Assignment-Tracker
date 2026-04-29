import tkinter as tk
from tkinter import ttk
from tkinter import messagebox

# Installing the ability to save assignments to excel file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create the main application window
window = tk.Tk()
window.title("Task/Assignment Tracker")
window.geometry("500x850")
window.configure(bg="#1e1e2e")

# Create the function to setup Excel file
def setup_excel():
    try:
        # Try to open existing file
        workbook = openpyxl.load_workbook("assignments.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        # File does not exist yet so create it
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Assignments"

        # Create header row
        headers = ["Assignment Name", "Class", "Type", "Due Date", "Target Date", "Priority", "Notes"]

        # Add headers to the first row
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header

            # Make headers bold and white text
            cell.font = Font(bold=True, color="FFFFFF")

            # Blue background for headers
            cell.fill = PatternFill("solid", fgColor="4472C4")

            # Center the text
            cell.alignment = Alignment(horizontal="center")
        
        # Set column widths so nothing gets cut off
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 25

        # Freeze the top row so the headers stay visible while scrolling
        sheet.freeze_panes = "A2"

        #Add filter dropdowns to the header row
        sheet.auto_filter.ref = "A1:G1"

        workbook.save("assignments.xlsx")
    return workbook, sheet

# Create a function to save an assignment to the Excel file
def save_to_excel(name, class_name, type_name, due, target, priority, notes):
    workbook = openpyxl.load_workbook("assignments.xlsx")
    sheet = workbook.active

    # Add the new row at the end
    row = [name, class_name, type_name, due, target, priority, notes]
    sheet.append(row)

    #Find the row we just added
    last_row = sheet.max_row

    # Set color based on priority
    if priority == "High":
        color = "FF6B6B"
    elif priority == "Medium":
        color = "F5A623"
    else:
        color = "4ECBA1"
    
    # Apply the color to the entire row
    for col in range(1, 7):
        cell = sheet.cell(row=last_row, column=col)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center")
    
    workbook.save("assignments.xlsx")

# Create a function to load assignments from the Excel file and display them in the listbox when the application starts
def load_assignments():
    try:
        workbook = openpyxl.load_workbook("assignments.xlsx")
        sheet = workbook.active

        # Skip the header row and load each assignment into the listbox
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Only add the row if it has data
            if row[0] is not None:
                name = row[0]
                class_name = row[1]
                type_name = row[2]
                due = row[3]
                target = row[4]
                priority = row[5]
                notes = row[6]
            assignment_list.insert(tk.END, f"{name} | {class_name} | {type_name} | Due: {due} | Target: {target} | {priority}")

        # Update the counter to match how many assignments are in the file
        global counter
        counter = sheet.max_row - 1
        counter_label.config(text=f"Assignments Added: {counter}")

    except FileNotFoundError:
        # If the file doesn't exist, just start with an empty list
        pass

# Create a function to delete the selected assignment from the listbox and the Excel file
def delete_assignment():
    # Check if something is selected in the Listbox
    selected = assignment_list.curselection()

    if not selected:
        messagebox.showerror("Error", "Please select an assignment to delete.")
        return
    
    # Get the index of the selected item
    index = selected[0]
    excel_row = index + 2  # +2 because Listbox is 0-indexed and Excel has a header row

    # Ask the user to confirm deletion
    confirm = messagebox.askyesno("Confirm Deletion", "Are you sure you want to delete the selected assignment?")

    if confirm:
        # Remove from Listbox
        assignment_list.delete(selected[0])

        # Remove from Excel
        workbook = openpyxl.load_workbook("assignments.xlsx")
        sheet = workbook.active
        sheet.delete_rows(excel_row)
        workbook.save("assignments.xlsx")

        # Update counter
        global counter
        counter -= 1
        counter_label.config(text=f"Assignments Added: {counter}")

        messagebox.showinfo("Deleted", "Assignment deleted successfully.")


# Counter variable to keep track of the number of assignments added
counter = 0

# Add checkbox variable
show_target = tk.IntVar()

# Call the function to set up the Excel file and get the workbook and sheet objects
workbook, sheet = setup_excel()

# Creating a placeholder text for the entry fields
def add_placeholder(entry, placeholder_text):
    entry.insert(0, placeholder_text)
    entry.config(fg='#888888')

    def on_click(event):
        if entry.get() == placeholder_text:
            entry.delete(0, tk.END)
            entry.config(fg='white')

    def on_leave(event):
        if entry.get() == '':
            entry.insert(0, placeholder_text)
            entry.config(fg='#888888')

    entry.bind("<FocusIn>", on_click)
    entry.bind("<FocusOut>", on_leave)

# Create the toggle function for the target completion date field
def toggle_target():
    if show_target.get() == 1:
        target_label.pack(after=target_check)
        target_entry.pack(after=target_label, pady=5)
        add_placeholder(target_entry, "e.g. 5/1/2026")
    else:
        target_label.pack_forget()
        target_entry.pack_forget()

# Create a function to make sure the date is valid
def is_valid_date(date_str):
    # Check it's not empty or placeholder
    if date_str == "" or date_str == "e.g. 4/27/2026":
        return False

    # Split by "/" and check we get exactly 3 parts
    parts = date_str.split("/")
    if len(parts) != 3:
        return False

    # Check each part is a number
    for part in parts:
        if not part.isdigit():
            return False

    # Pull out month, day, year
    month = int(parts[0])
    day   = int(parts[1])
    year  = int(parts[2])

    # Check ranges make sense
    if month < 1 or month > 12:
        return False
    if day < 1 or day > 31:
        return False
    if year < 2000 or year > 2100:
        return False

    return True
#Title at the top of the window
title_label = tk.Label(window, text="Task/Assignment Tracker", font=("Arial", 20, "bold"), bg="#1e1e2e", fg="#6c9FFF")
title_label.pack(pady=20)

# Assignment name
name_label = tk.Label(window, text="Assignment Name:", font=("Arial", 12), bg="#1e1e2e", fg="white")
name_label.pack()
name_entry = tk.Entry(window, font=("Arial", 12), width=30, bg="#2e2e3e", fg="white", insertbackground="white")
name_entry.pack(pady=5)
add_placeholder(name_entry, "e.g. Math Homework 5")

# Class
class_label = tk.Label(window, text="Class:", font=("Arial", 12), bg="#1e1e2e", fg="white")
class_label.pack()
class_entry = tk.Entry(window, font=("Arial", 12), width=30, bg="#2e2e3e", fg="white", insertbackground="white")
class_entry.pack(pady=5)
add_placeholder(class_entry, "e.g. Math 101")

# Due date
due_label = tk.Label(window, text="Due Date (MM/DD/YYYY):", font=("Arial", 12), bg="#1e1e2e", fg="white")
due_label.pack()
due_entry = tk.Entry(window, font=("Arial", 12), width=30, bg="#2e2e3e", fg="white", insertbackground="white")
due_entry.pack(pady=5)
add_placeholder(due_entry, "e.g. 4/27/2026")

#Type of assignment dropdown
type_label = tk.Label(window, text="Assignment Type:", font=("Arial", 12), bg="#1e1e2e", fg="white")
type_label.pack()
type_box = ttk.Combobox(window, values=["Homework", "Quiz", "Exam", "Project", "Lab", "Paper", "Reading", "Other"], font=("Arial", 12), width=28, state="readonly")
type_box.pack(pady=5)

# Priority dropdown
priority_label = tk.Label(window, text="Priority:", font=("Arial", 12), bg="#1e1e2e", fg="white")
priority_label.pack()
priority_box = ttk.Combobox(window, values=["Low", "Medium", "High"], font=("Arial", 12), width=28, state="readonly")
priority_box.pack(pady=5)

# Notes
notes_label = tk.Label(window, text="Notes (Optional):", font=("Arial", 12), bg="#1e1e2e", fg="white")
notes_label.pack()
notes_entry = tk.Entry(window, font=("Arial", 12), width=30, bg="#2e2e3e", fg="white", insertbackground="white")
notes_entry.pack(pady=5)
add_placeholder(notes_entry, "e.g. Chapters 5-7")

# Target completion date checkbox
target_check = tk.Checkbutton(window, text="Set Target Completion Date?", font=("Arial", 11), bg="#1e1e2e", fg="white", selectcolor="#2e2e3e", activebackground="#1e1e2e", activeforeground="white", variable=show_target, command=toggle_target)
target_check.pack(pady=5)

# Target date field (hidden by default)
target_label = tk.Label(window, text="Target Completion Date (MM/DD/YYYY):", font=("Arial", 12), bg="#1e1e2e", fg="white")
target_entry = tk.Entry(window, font=("Arial", 12), width=30, bg="#2e2e3e", fg="white", insertbackground="white")
# Submit function
def submit():
    global counter

    name = name_entry.get()
    class_name = class_entry.get()
    due = due_entry.get()
    priority = priority_box.get()
    notes = notes_entry.get()
    if notes == "e.g. Chapters 5-7":
        notes = ""
    type_name = type_box.get()

    if show_target.get() == 1:
        target = target_entry.get()
        if target == "e.g. 5/1/2026":
            target = ""
    else:
        target = ""

    # Make sure nothing is empty
    if name == "" or name == "e.g. Math Homework 5" or class_name == "" or class_name == "e.g. Math 101" or due == "" or due == "e.g. 4/27/2026" or priority == "" or type_name == "":
        messagebox.showerror("Error", "Please fill in all fields.")
    elif not is_valid_date(due):
        messagebox.showerror("Invalid Date", "Please enter a valid date in MM/DD/YYYY format.")
    elif show_target.get() == 1 and not is_valid_date(target_entry.get()):
        messagebox.showerror("Invalid Target Date", "Please enter a valid target completion date in MM/DD/YYYY format.")
    else:
        messagebox.showinfo("Success", f"Added: {name}\nType: {type_name}\nNotes: {notes if notes else 'None'}")

        # Save the assignment to the Excel file
        save_to_excel(name, class_name, type_name, due, target, priority, notes)

        # Add the assignment to the listbox
        assignment_list.insert(tk.END, f"{name} | {class_name} | {type_name} | Due: {due} | Target: {target} | {priority}")

        # Clear the fields after submission
        name_entry.delete(0, tk.END)
        class_entry.delete(0, tk.END)
        due_entry.delete(0, tk.END)
        priority_box.set("")
        type_box.set("")
        notes_entry.delete(0, tk.END)

        if show_target.get() == 1:
            target_entry.delete(0, tk.END)
            target_check.deselect()
            target_label.pack_forget()
            target_entry.pack_forget()

        # Increment the counter and update the label
        counter += 1
        counter_label.config(text=f"Assignments Added: {counter}")

# Counter label
counter_label = tk.Label(window, text=f"Assignments Added: 0", font=("Arial", 11), bg="#1e1e2e", fg="gray")
counter_label.pack(pady=5)

# Submit button
submit_button = tk.Button(window, text="Submit", font=("Arial", 12, "bold"), bg="#6c9FFF", fg="black", activebackground="#4a7fd4", activeforeground="white", relief="flat", padx=20, pady=8, command=submit)
submit_button.pack(pady=20)

# List to display added assignments
list_label = tk.Label(window, text="Assignments This Session:", font=("Arial", 12, "bold"), bg="#1e1e2e", fg="white")
list_label.pack(pady=(10, 0))

# Frame to hold the listbox and scrollbar
list_frame = tk.Frame(window, bg="#1e1e2e")
list_frame.pack(pady=5)

# Scrollbar for the listbox
scrollbar = tk.Scrollbar(list_frame)
scrollbar.pack(side="right", fill="y")

# Listbox to display assignments
assignment_list = tk.Listbox(list_frame, font=("Arial", 11), width=50, height=6, bg="#2e2e3e", fg="white", selectbackground="#6c9FFF", yscrollcommand=scrollbar.set, bd=0)
assignment_list.pack(side="left")

# Connect scrollbar to listbox
scrollbar.config(command=assignment_list.yview)

# Delete button
delete_button = tk.Button(window, text="Delete Selected", font=("Arial", 11, "bold"), bg="#FF6B6B", fg="black", activebackground="#cc0000", activeforeground="white", relief="flat", padx=10, pady=6, command=delete_assignment)
delete_button.pack(pady=5)

# Load existing assignments from the Excel file when the application starts
load_assignments()

# Start the window
window.mainloop()



